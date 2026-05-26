"""
Seed script: loads all medicine and symptom data from the provided Excel files
into the database.

Usage (from the ProjectFixed folder):
    python scripts/seed_data.py [--force]

Options:
    --force     Bypass the idempotency guard and re-import even if medicines exist.

Files expected inside checkfordata/ (project root subfolder):
  - checkfordata/All_A_Medicines_With_Food_Interactions.xlsx   (standalone A)
  - checkfordata/All_B_Medicines_With_Food_Interactions.xlsx   (standalone B)
  - checkfordata/All_C_Medicines_With_Food_Interactions.xlsx   (standalone C)
  - checkfordata/ab.xlsx                                       (dosage data A-Z)
  - checkfordata/drug-drug-interactions-formatted.xlsx         (drug interactions)
  - workspace-*.zip                                            (D-Z letters, at root)
  - symptoms.xlsx                                              (symptom rules, at root)

Idempotency: if the database already contains more than IDEMPOTENCY_THRESHOLD
medicines, the script exits early, so it is safe to run on every container
start without re-seeding.  Use --force to bypass.

FK Safety: an ImportBatch row is created and committed BEFORE any Medicine rows
are inserted.  The committed batch.id is used as medicine.import_batch_id,
preserving the foreign key constraint on every DB engine (including PostgreSQL).
"""

import argparse
import io
import logging
import sys
import uuid
import zipfile
from pathlib import Path

# ---------------------------------------------------------------------------
# Allow running from project root: python scripts/seed_data.py
# ---------------------------------------------------------------------------
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

# Configure logging so output is visible in Docker/Render logs
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [seed_data] %(message)s",
)
logger = logging.getLogger(__name__)

from openpyxl import load_workbook
from sqlalchemy import func, select as sa_select
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.models.medicine import ImportBatch, Medicine, MedicineAlias, MedicineSource
from app.models.symptom import SymptomRule

# ---------------------------------------------------------------------------
# File paths — edit these if your files are stored elsewhere
# ---------------------------------------------------------------------------
BASE = _ROOT
CHECKFORDATA = BASE / "checkfordata"

# Medicine catalog files — now read from checkfordata/ subfolder
DATA_FILES_ABC = [
    CHECKFORDATA / "All_A_Medicines_With_Food_Interactions.xlsx",
    CHECKFORDATA / "All_B_Medicines_With_Food_Interactions.xlsx",
    CHECKFORDATA / "All_C_Medicines_With_Food_Interactions.xlsx",
]
# ZIP file for D-Z medicines — still expected at project root (original location)
ZIP_FILE = next(BASE.glob("workspace-*.zip"), None)
# Dosage reference — now read from checkfordata/ subfolder
AB_DOSAGE_FILE = CHECKFORDATA / "ab.xlsx"
# Symptoms file — remains at project root (only copy)
SYMPTOMS_FILE = BASE / "symptoms.xlsx"

SOURCE_NAME = "drugs.com"

# Threshold: if the DB already has more than this many medicines, skip seeding
IDEMPOTENCY_THRESHOLD = 10


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def new_id() -> str:
    return str(uuid.uuid4())


def clean(val) -> str | None:
    """Strip and return None for empty/null values."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("none", "n/a", "na", "-", "") else None


def parse_aliases(indian_brands: str | None) -> list[str]:
    """Split 'Brand1, Brand2 (import)' into individual alias strings."""
    if not indian_brands:
        return []
    parts = [p.strip() for p in indian_brands.replace(";", ",").split(",")]
    return [p for p in parts if p and p.lower() not in ("same", "not common", "not available",
                                                          "not widely available", "limited availability")]


# ---------------------------------------------------------------------------
# 1. Load medicine data from A/B/C standalone files
# ---------------------------------------------------------------------------

def load_abc_file(path: Path) -> list[dict]:
    """
    Columns (row 2 is header):
    Brand Name | Generic Name | Indian Brands | Drug Class |
    Symptoms / When to Take | Diseases / Medical Conditions Treated |
    Sources | Foods to Avoid / Dietary Interactions
    """
    wb = load_workbook(path, read_only=True)
    # Pick the main data sheet (first one with 'Medicine' in the name or Sheet1)
    ws = None
    for sh in wb.sheetnames:
        if "medicine" in sh.lower() or sh.lower().startswith("sheet"):
            ws = wb[sh]
            break
    if ws is None:
        ws = wb.active

    records = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            # Skip until we hit the header row
            if row[0] and str(row[0]).strip().lower().startswith("brand"):
                header_found = True
            continue
        if not row[0]:
            continue
        brand, generic, indian, drug_class, symptoms, diseases, sources, foods = (row + (None,) * 8)[:8]
        if not clean(generic):
            continue
        records.append({
            "brand_name": clean(brand),
            "generic_name": clean(generic),
            "indian_brands": clean(indian),
            "drug_class": clean(drug_class),
            "symptoms": clean(symptoms),
            "diseases": clean(diseases),
            "sources": clean(sources),
            "foods_to_avoid": clean(foods),
        })
    return records


# ---------------------------------------------------------------------------
# 2. Load medicine data from ZIP (D–Z)
# ---------------------------------------------------------------------------

def load_zip_outputs(zip_path: Path) -> list[dict]:
    """
    Columns (row 1 is header):
    Brand Name | Generic Name | Indian Brands | Drug Class/Category |
    Primary Uses | Key Side Effects | Contraindications | Foods to Avoid
    """
    records = []
    with zipfile.ZipFile(zip_path) as z:
        outputs = [n for n in z.namelist() if n.startswith("outputs/") and n.endswith(".xlsx")]
        for name in outputs:
            with z.open(name) as f:
                wb = load_workbook(io.BytesIO(f.read()), read_only=True)
                ws = wb.active
                header_found = False
                for row in ws.iter_rows(values_only=True):
                    if not header_found:
                        if row[0] and str(row[0]).strip().lower().startswith("brand"):
                            header_found = True
                        continue
                    if not row[0]:
                        continue
                    brand, generic, indian, drug_class, primary_uses, side_effects, contraindications, foods = (row + (None,) * 8)[:8]
                    if not clean(generic):
                        continue
                    records.append({
                        "brand_name": clean(brand),
                        "generic_name": clean(generic),
                        "indian_brands": clean(indian),
                        "drug_class": clean(drug_class),
                        "symptoms": clean(primary_uses),
                        "diseases": clean(primary_uses),
                        "side_effects": clean(side_effects),
                        "contraindications": clean(contraindications),
                        "foods_to_avoid": clean(foods),
                        "sources": None,
                    })
    return records


# ---------------------------------------------------------------------------
# 3. Load dosage data from ab.xlsx
# ---------------------------------------------------------------------------

def load_dosage_map(path: Path) -> dict[str, dict]:
    """Returns a dict keyed by lowercase generic_name -> dosage info."""
    dosage_map: dict[str, dict] = {}
    wb = load_workbook(path, read_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if not header_found:
                if row[0] and str(row[0]).strip().lower().startswith("brand"):
                    header_found = True
                continue
            if not row[0]:
                continue
            # Columns: Brand, Generic, Indian Brands, Infant, Pediatric, Adult, Geriatric, Common Dosage, ...
            cols = list(row) + [None] * 16
            generic     = clean(cols[1])
            adult_dose  = clean(cols[5])
            common_dose = clean(cols[7])
            if not generic:
                continue
            key = generic.lower()
            dosage_map[key] = {
                "adult_dose": adult_dose,
                "common_dose": common_dose,
            }
    return dosage_map


# ---------------------------------------------------------------------------
# 4. Load symptom rules from symptoms.xlsx
# ---------------------------------------------------------------------------

def load_symptom_rules(path: Path) -> list[dict]:
    """
    Columns (row 2 is header):
    Brand Name | Generic Name | Indian Brands | Drug Class |
    Symptoms / When to Take | Diseases / Medical Conditions Treated | Sources
    We derive SymptomRule from: symptoms -> keywords, diseases -> condition, generic -> medicine
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rules = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] and str(row[0]).strip().lower().startswith("brand"):
                header_found = True
            continue
        if not row[0]:
            continue
        cols = list(row) + [None] * 8
        generic  = clean(cols[1])
        symptoms = clean(cols[4])
        diseases = clean(cols[5])
        if not generic or not symptoms:
            continue

        # Build keyword list from symptom text (split on ; and ,)
        import re
        raw_kws = re.split(r"[;,]", symptoms)
        keywords = [kw.strip().lower() for kw in raw_kws if len(kw.strip()) > 3][:10]

        # One rule per condition listed
        raw_conditions = re.split(r"[;,]", diseases or generic)
        for condition in raw_conditions[:1]:  # one rule per medicine to avoid duplicates
            condition = condition.strip()
            if not condition:
                continue
            rules.append({
                "name": f"{condition} — {generic}",
                "symptom_keywords": "|".join(keywords),
                "possible_condition": condition[:255],
                "care_recommendations": f"Consult a doctor. Commonly managed with {generic}.",
                "common_medicines": generic,
                "precautions": "Do not self-medicate. Seek professional advice.",
                "escalation_triggers": "severe symptoms|difficulty breathing|chest pain|loss of consciousness",
                "is_emergency": False,
            })
    return rules


# ---------------------------------------------------------------------------
# 5. Create the ImportBatch row (commit BEFORE inserting medicines)
# ---------------------------------------------------------------------------

def _repair_orphan_batch_refs(db: Session) -> None:
    """
    Self-healing repair: find medicines whose import_batch_id references
    a non-existent import_batches row (FK violation) and remap them to
    the latest completed batch.  If no completed batch exists, create one.

    This runs on every start of seed_data.py so existing deployments that
    ran the old seed (which never inserted an ImportBatch) are healed
    automatically — even when the idempotency guard skips the full seed.
    """
    from sqlalchemy import text

    orphan_count = db.execute(text("""
        SELECT COUNT(*) FROM medicines m
        LEFT JOIN import_batches b ON m.import_batch_id = b.id
        WHERE m.import_batch_id IS NOT NULL AND b.id IS NULL
    """)).scalar() or 0

    null_count = db.execute(text(
        "SELECT COUNT(*) FROM medicines WHERE import_batch_id IS NULL"
    )).scalar() or 0

    total_broken = orphan_count + null_count
    if total_broken == 0:
        logger.debug("FK check: all medicine.import_batch_id values are valid.")
        return

    logger.warning(
        "Detected %d medicines with broken import_batch_id FK references "
        "(%d orphans, %d NULL). Repairing ...",
        total_broken, orphan_count, null_count,
    )

    # Find or create a valid target batch
    target = db.execute(text(
        "SELECT id FROM import_batches WHERE status = 'completed' ORDER BY created_at DESC LIMIT 1"
    )).fetchone()

    if target:
        target_id = target[0]
        logger.info("Remapping broken references to existing batch %s ...", target_id[:8])
    else:
        # No completed batch exists at all — create a repair batch
        repair_batch = ImportBatch(
            id=new_id(),
            source_name=SOURCE_NAME,
            source_type="repair-seed",
            filename="legacy-seed-repair",
            status="completed",
            records_total=0,
            records_imported=0,
        )
        db.add(repair_batch)
        db.commit()
        db.refresh(repair_batch)
        target_id = repair_batch.id
        logger.info("Created repair batch %s.", target_id[:8])

    # Remap all broken FK references
    result = db.execute(text("""
        UPDATE medicines
        SET import_batch_id = :bid
        WHERE import_batch_id NOT IN (SELECT id FROM import_batches)
           OR import_batch_id IS NULL
    """), {"bid": target_id})
    db.commit()
    logger.info("Repaired %d medicine rows → batch %s.", result.rowcount, target_id[:8])

    # Update batch record count
    db.execute(text("""
        UPDATE import_batches
        SET records_imported = (
            SELECT COUNT(*) FROM medicines WHERE import_batch_id = :bid
        )
        WHERE id = :bid
    """), {"bid": target_id})
    db.commit()


def create_import_batch(db: Session, records_total: int) -> ImportBatch:
    """
    Insert and COMMIT an ImportBatch row so its ID exists in PostgreSQL
    before any Medicine FK references it.

    Returns the refreshed ImportBatch instance with a valid .id.
    """
    batch = ImportBatch(
        id=new_id(),
        source_name=SOURCE_NAME,
        source_type="xlsx-seed",
        filename="All_A/B/C_Medicines + ab.xlsx + workspace-*.zip",
        status="pending",
        records_total=records_total,
        records_imported=0,
    )
    db.add(batch)
    db.commit()       # ← FK-safe: the row now exists in PostgreSQL
    db.refresh(batch) # ← ensures batch.id is populated from DB
    logger.info("ImportBatch created: id=%s records_total=%d", batch.id, records_total)
    return batch


# ---------------------------------------------------------------------------
# 6. Upsert medicines into DB (using committed batch.id)
# ---------------------------------------------------------------------------

def upsert_medicines(
    db: Session,
    records: list[dict],
    dosage_map: dict[str, dict],
    batch_id: str,
) -> tuple[int, int]:
    inserted = 0
    skipped = 0
    for rec in records:
        generic = rec.get("generic_name")
        if not generic:
            continue

        # Check duplicate by generic_name (case-insensitive)
        existing = db.query(Medicine).filter(
            Medicine.generic_name.ilike(generic)
        ).first()

        dos = dosage_map.get(generic.lower(), {})

        # Build usage_guidelines from symptoms + diseases + dosage (NOT foods)
        usage_parts = []
        if rec.get("symptoms"):
            usage_parts.append(f"Symptoms: {rec['symptoms']}")
        if rec.get("diseases"):
            usage_parts.append(f"Conditions: {rec['diseases']}")
        if dos.get("common_dose"):
            usage_parts.append(f"Common dosage: {dos['common_dose']}")
        if dos.get("adult_dose"):
            usage_parts.append(f"Adult dose: {dos['adult_dose']}")
        usage_guidelines = "\n".join(usage_parts) if usage_parts else None

        # Foods to avoid stored in the dedicated precautions field
        precautions_text = rec.get("foods_to_avoid")

        # Build dosage_form / strength from drug_class
        drug_class = rec.get("drug_class")

        if existing:
            # Enrich existing record with missing data
            if not existing.usage_guidelines and usage_guidelines:
                existing.usage_guidelines = usage_guidelines
            if not existing.precautions and precautions_text:
                existing.precautions = precautions_text
            if not existing.side_effects and rec.get("side_effects"):
                existing.side_effects = rec["side_effects"]
            if not existing.contraindications and rec.get("contraindications"):
                existing.contraindications = rec["contraindications"]
            if not existing.composition and drug_class:
                existing.composition = drug_class
            # Backfill import_batch_id on enriched rows that have no batch yet
            if not existing.import_batch_id:
                existing.import_batch_id = batch_id
            skipped += 1
            med = existing
        else:
            med = Medicine(
                id=new_id(),
                generic_name=generic,
                brand_name=rec.get("brand_name"),
                composition=drug_class,
                side_effects=rec.get("side_effects"),
                contraindications=rec.get("contraindications"),
                precautions=precautions_text,
                usage_guidelines=usage_guidelines,
                source_name=SOURCE_NAME,
                import_batch_id=batch_id,   # ← uses committed batch.id
            )
            db.add(med)
            inserted += 1

        db.flush()

        # Add Indian brand aliases
        existing_aliases = {a.alias.lower() for a in db.query(MedicineAlias).filter_by(medicine_id=med.id).all()}
        for alias in parse_aliases(rec.get("indian_brands")):
            if alias.lower() not in existing_aliases:
                db.add(MedicineAlias(
                    id=new_id(),
                    medicine_id=med.id,
                    alias=alias,
                    alias_type="indian_brand",
                    source_name=SOURCE_NAME,
                ))
                existing_aliases.add(alias.lower())

        # Add brand name as alias too (if different from generic)
        brand = rec.get("brand_name")
        if brand and brand.lower() != generic.lower() and brand.lower() not in existing_aliases:
            db.add(MedicineAlias(
                id=new_id(),
                medicine_id=med.id,
                alias=brand,
                alias_type="brand",
                source_name=SOURCE_NAME,
            ))

    return inserted, skipped


# ---------------------------------------------------------------------------
# 7. Insert symptom rules
# ---------------------------------------------------------------------------

def insert_symptom_rules(db: Session, rules: list[dict]) -> int:
    inserted = 0
    existing_names = {r.name for r in db.query(SymptomRule.name).all()}
    for rule in rules:
        if rule["name"] in existing_names:
            continue
        db.add(SymptomRule(
            id=new_id(),
            name=rule["name"],
            symptom_keywords=rule["symptom_keywords"],
            possible_condition=rule["possible_condition"],
            care_recommendations=rule["care_recommendations"],
            common_medicines=rule["common_medicines"],
            precautions=rule["precautions"],
            escalation_triggers=rule["escalation_triggers"],
            is_emergency=rule["is_emergency"],
        ))
        existing_names.add(rule["name"])
        inserted += 1
    return inserted


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Seed medicine catalog into the database.")
    parser.add_argument("--force", action="store_true", help="Bypass idempotency guard and re-import")
    args = parser.parse_args()

    logger.info("=" * 60)
    logger.info("Smart Drug Safety — Full Data Seeder")
    logger.info("=" * 60)

    db: Session = SessionLocal()
    batch: ImportBatch | None = None

    try:
        # ── FK repair: heal any orphan import_batch_id references ─────────
        # Runs on EVERY invocation — also when the idempotency guard triggers.
        # Fixes existing deployments where medicines were seeded without a
        # corresponding ImportBatch row (the old seed bug).
        _repair_orphan_batch_refs(db)

        # ── Idempotency check ─────────────────────────────────────
        existing_count = db.scalar(sa_select(func.count()).select_from(Medicine)) or 0
        if existing_count > IDEMPOTENCY_THRESHOLD and not args.force:
            logger.info(
                "DB already contains %d medicines (threshold=%d). "
                "Skipping full seed — idempotency guard triggered.",
                existing_count,
                IDEMPOTENCY_THRESHOLD,
            )
            return
        logger.info("DB has %d medicines. Proceeding with full seed.", existing_count)

        # ── Load all source files first (before any DB writes) ────
        logger.info("[1/4] Loading dosage map from ab.xlsx ...")
        if AB_DOSAGE_FILE.exists():
            dosage_map = load_dosage_map(AB_DOSAGE_FILE)
            logger.info("      %d dosage entries loaded.", len(dosage_map))
        else:
            logger.warning("      %s not found. Dosage data skipped.", AB_DOSAGE_FILE)
            dosage_map = {}

        logger.info("[2/4] Loading medicine records from standalone A/B/C files ...")
        abc_records: list[dict] = []
        for f in DATA_FILES_ABC:
            if f.exists():
                try:
                    recs = load_abc_file(f)
                    logger.info("      %s: %d medicines", f.name, len(recs))
                    abc_records.extend(recs)
                except Exception as exc:
                    logger.warning("      Skipping %s due to error: %s", f.name, exc)
            else:
                logger.warning("      %s not found — skipped.", f.name)

        logger.info("[3/4] Loading medicine records from ZIP (D-Z) ...")
        zip_records: list[dict] = []
        if ZIP_FILE and ZIP_FILE.exists():
            try:
                zip_records = load_zip_outputs(ZIP_FILE)
                logger.info("      %d medicines loaded from ZIP.", len(zip_records))
            except Exception as exc:
                logger.warning("      ZIP load failed: %s — skipped.", exc)
        else:
            logger.warning("      ZIP file not found — skipped.")

        all_records = abc_records + zip_records
        logger.info("      Total records to process: %d", len(all_records))

        # ── Create ImportBatch FIRST and commit it ─────────────────
        # This ensures the FK import_batches.id exists in PostgreSQL
        # before any Medicine row references it.
        batch = create_import_batch(db, records_total=len(all_records))

        # ── Insert medicines using the committed batch.id ──────────
        ins, skip = upsert_medicines(db, all_records, dosage_map, batch_id=batch.id)
        db.commit()
        logger.info("      Inserted: %d  |  Enriched existing: %d", ins, skip)

        # ── Mark batch as completed ────────────────────────────────
        batch.status = "completed"
        batch.records_imported = ins
        db.commit()
        logger.info("ImportBatch %s marked completed (%d imported).", batch.id, ins)

        # ── Symptom Rules ──────────────────────────────────────────
        logger.info("[4/4] Loading symptom rules from symptoms.xlsx ...")
        if SYMPTOMS_FILE.exists():
            try:
                rules = load_symptom_rules(SYMPTOMS_FILE)
                logger.info("      %d rules parsed.", len(rules))
                rule_count = insert_symptom_rules(db, rules)
                db.commit()
                logger.info("      Inserted: %d symptom rules.", rule_count)
            except Exception as exc:
                logger.warning("      Symptom rules load failed: %s — skipped.", exc)
        else:
            logger.warning("      %s not found — skipped.", SYMPTOMS_FILE)

        # ── Summary ────────────────────────────────────────────────
        total_med = db.scalar(sa_select(func.count()).select_from(Medicine)) or 0
        total_alias = db.query(MedicineAlias).count()
        total_sym = db.query(SymptomRule).count()
        logger.info("=" * 60)
        logger.info("  Medicines in DB : %d", total_med)
        logger.info("  Aliases in DB   : %d", total_alias)
        logger.info("  Symptom rules   : %d", total_sym)
        logger.info("=" * 60)
        logger.info("  Seeding complete!")

    except Exception as e:
        db.rollback()
        # Mark the batch as failed if it was created
        if batch is not None:
            try:
                batch.status = "failed"
                batch.errors = str(e)[:500]
                db.commit()
                logger.error("ImportBatch %s marked as failed.", batch.id)
            except Exception:
                pass  # Don't mask the original error
        logger.exception("Seeding failed: %s", e)
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
