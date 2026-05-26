"""
Seed script: imports drug-drug interactions from checkfordata/drug-drug-interactions-formatted.xlsx
into the drug_interactions table.

Usage (from the ProjectFixed folder):
    python scripts/seed_interactions.py [--dry-run] [--force]

Options:
    --dry-run   Parse and validate the file but do NOT write to the database.
    --force     Skip the idempotency guard and re-import even if interactions exist.

Expected columns in the xlsx (row 1 = header):
    Drug 1 | Drug 2 | Interaction Description | Severity Level |
    Safety Label | Drug Class 1 | Drug Class 2 | Diseases Treated

Severity mapping:
    High     -> "high"
    Moderate -> "moderate"
    (others) -> lowercased as-is

Safety Label -> stored in recommendations field:
    NOT SAFE             -> "NOT SAFE — avoid concurrent use."
    SAFE WITH MONITORING -> "SAFE WITH MONITORING — use with caution and monitor closely."
"""

import argparse
import logging
import re
import sys
import uuid
from pathlib import Path

# ---------------------------------------------------------------------------
# Ensure project root is on sys.path so app.* imports work
# ---------------------------------------------------------------------------
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [seed_interactions] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.models.interaction import DrugInteraction

# ---------------------------------------------------------------------------
# Path configuration — reads from checkfordata/ inside the project root
# ---------------------------------------------------------------------------
INTERACTIONS_FILE = _ROOT / "checkfordata" / "drug-drug-interactions-formatted.xlsx"

SOURCE_NAME = "drugs.com-formatted"
IDEMPOTENCY_THRESHOLD = 10  # if DB already has this many, skip unless --force

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def new_id() -> str:
    return str(uuid.uuid4())


def _clean(val) -> str | None:
    """Strip whitespace and return None for empty/null-like values."""
    if val is None:
        return None
    text = str(val).strip()
    return text if text and text.lower() not in {"none", "n/a", "na", "-", ""} else None


def _normalize_key(value: str) -> str:
    """Lowercase, collapse whitespace, strip punctuation for dedup key."""
    if not value:
        return ""
    key = value.strip().casefold()
    key = re.sub(r"[^\w\s]", "", key)   # remove punctuation
    key = re.sub(r"\s+", " ", key).strip()
    return key


def _map_severity(raw: str | None) -> str:
    """Normalize severity to a consistent string."""
    if not raw:
        return "unknown"
    raw_lower = raw.strip().lower()
    if raw_lower in {"high", "severe", "major"}:
        return "high"
    if raw_lower in {"moderate", "medium"}:
        return "moderate"
    if raw_lower in {"low", "minor", "minimal"}:
        return "low"
    return raw_lower  # preserve unknown values as-is


def _map_safety_label(raw: str | None) -> str | None:
    """Convert safety label to a human-readable recommendation string."""
    if not raw:
        return None
    label = raw.strip().upper()
    if "NOT SAFE" in label:
        return "NOT SAFE — avoid concurrent use without specialist supervision."
    if "MONITORING" in label or "CAUTION" in label:
        return "SAFE WITH MONITORING — use with caution and monitor closely for adverse effects."
    return raw.strip()


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def parse_interactions_xlsx(path: Path) -> list[dict]:
    """
    Parse the drug-drug interactions xlsx file.

    Returns a list of dicts with keys:
        drug_a_name, drug_b_name, drug_a_key, drug_b_key,
        severity, explanation, mechanism, recommendations, source_name
    """
    if not path.exists():
        logger.error("Interactions file not found: %s", path)
        return []

    logger.info("Parsing: %s", path)
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    records: list[dict] = []
    skipped_rows = 0
    header_found = False

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Detect header row dynamically (look for "Drug 1" in first cell)
        if not header_found:
            if row[0] and str(row[0]).strip().lower() in {"drug 1", "drug1", "drug_1", "drug a"}:
                header_found = True
                logger.debug("Header found at row %d: %s", row_idx, list(row))
            continue

        # Skip blank rows
        if not row[0] and not row[1]:
            skipped_rows += 1
            continue

        # Unpack columns (pad to 8 in case of short rows)
        cols = list(row) + [None] * 8
        drug_1          = _clean(cols[0])
        drug_2          = _clean(cols[1])
        interaction_desc = _clean(cols[2])
        severity_raw    = _clean(cols[3])
        safety_label    = _clean(cols[4])
        drug_class_1    = _clean(cols[5])
        drug_class_2    = _clean(cols[6])
        diseases        = _clean(cols[7])

        # Both drug names are required
        if not drug_1 or not drug_2:
            logger.debug("Row %d skipped — missing drug name(s): drug_1=%r drug_2=%r", row_idx, drug_1, drug_2)
            skipped_rows += 1
            continue

        # Derive canonical sort keys (always store pair in alphabetical order)
        key_a = _normalize_key(drug_1)
        key_b = _normalize_key(drug_2)
        if key_a > key_b:
            key_a, key_b = key_b, key_a
            drug_1, drug_2 = drug_2, drug_1
            # Also swap classes to stay consistent
            drug_class_1, drug_class_2 = drug_class_2, drug_class_1

        # Build mechanism from drug classes if available
        mechanism_parts = []
        if drug_class_1:
            mechanism_parts.append(f"{drug_1}: {drug_class_1}")
        if drug_class_2:
            mechanism_parts.append(f"{drug_2}: {drug_class_2}")
        if diseases:
            mechanism_parts.append(f"Conditions treated: {diseases}")
        mechanism = "; ".join(mechanism_parts) if mechanism_parts else None

        records.append({
            "drug_a_name": drug_1,
            "drug_b_name": drug_2,
            "drug_a_key": key_a,
            "drug_b_key": key_b,
            "severity": _map_severity(severity_raw),
            "explanation": interaction_desc or f"Interaction between {drug_1} and {drug_2}.",
            "mechanism": mechanism,
            "recommendations": _map_safety_label(safety_label),
            "source_name": SOURCE_NAME,
            "source_url": None,
            "source_version": "2024",
            "confidence": 0.9,
        })

    if not header_found:
        logger.error("Header row not found in %s — check the file format.", path)

    logger.info("Parsed %d interaction records (%d rows skipped).", len(records), skipped_rows)
    return records


# ---------------------------------------------------------------------------
# Database insertion
# ---------------------------------------------------------------------------

def seed_interactions(db: Session, records: list[dict], dry_run: bool = False) -> dict:
    """
    Upsert interaction records into drug_interactions table.

    Returns:
        dict with keys: inserted, updated, skipped, failed
    """
    inserted = 0
    updated = 0
    skipped = 0
    failed = 0
    seen_pairs: set[tuple[str, str]] = set()  # deduplicate within this batch

    for rec in records:
        key_a = rec["drug_a_key"]
        key_b = rec["drug_b_key"]
        pair = (key_a, key_b)

        # Skip within-batch duplicates
        if pair in seen_pairs:
            logger.debug("Batch duplicate skipped: %s <-> %s", rec["drug_a_name"], rec["drug_b_name"])
            skipped += 1
            continue
        seen_pairs.add(pair)

        try:
            existing = db.scalar(
                select(DrugInteraction).where(
                    DrugInteraction.drug_a_key == key_a,
                    DrugInteraction.drug_b_key == key_b,
                )
            )

            if existing:
                # Enrich existing record with any missing fields
                changed = False
                for field in ("explanation", "mechanism", "recommendations", "severity", "source_name"):
                    new_val = rec.get(field)
                    if new_val and not getattr(existing, field):
                        setattr(existing, field, new_val)
                        changed = True
                if changed:
                    updated += 1
                    logger.debug("Updated: %s <-> %s", rec["drug_a_name"], rec["drug_b_name"])
                else:
                    skipped += 1
                continue

            if not dry_run:
                db.add(
                    DrugInteraction(
                        id=new_id(),
                        drug_a_key=key_a,
                        drug_b_key=key_b,
                        drug_a_name=rec["drug_a_name"],
                        drug_b_name=rec["drug_b_name"],
                        severity=rec["severity"],
                        explanation=rec["explanation"],
                        mechanism=rec["mechanism"],
                        recommendations=rec["recommendations"],
                        source_name=rec["source_name"],
                        source_url=rec["source_url"],
                        source_version=rec["source_version"],
                        confidence=rec["confidence"],
                    )
                )
            inserted += 1
            logger.debug("Inserting: %s <-> %s [%s]", rec["drug_a_name"], rec["drug_b_name"], rec["severity"])

        except Exception as exc:
            logger.warning(
                "Failed to process interaction %s <-> %s: %s",
                rec.get("drug_a_name", "?"), rec.get("drug_b_name", "?"), exc,
            )
            failed += 1
            continue

        # Flush periodically to keep memory usage low
        if not dry_run and inserted % 50 == 0 and inserted > 0:
            db.flush()

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "failed": failed}


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Seed drug-drug interactions from checkfordata/")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, do not write to DB")
    parser.add_argument("--force", action="store_true", help="Bypass idempotency guard")
    args = parser.parse_args()

    logger.info("=" * 60)
    logger.info("Smart Drug Safety — Interactions Seeder")
    if args.dry_run:
        logger.info("  Mode: DRY RUN (no DB writes)")
    logger.info("=" * 60)

    # ── Parse file ────────────────────────────────────────────────
    records = parse_interactions_xlsx(INTERACTIONS_FILE)
    if not records:
        logger.error("No records parsed. Aborting.")
        sys.exit(1)

    logger.info("Total interaction records to process: %d", len(records))

    if args.dry_run:
        logger.info("DRY RUN complete. Would insert up to %d interactions.", len(records))
        logger.info("Severity breakdown:")
        from collections import Counter
        severities = Counter(r["severity"] for r in records)
        for sev, count in sorted(severities.items()):
            logger.info("  %-20s %d", sev, count)
        return

    # ── Database operations ───────────────────────────────────────
    db: Session = SessionLocal()
    try:
        # Idempotency check
        existing_count = db.scalar(
            select(func.count()).select_from(DrugInteraction)
        ) or 0
        logger.info("Current interaction count in DB: %d", existing_count)

        if existing_count >= IDEMPOTENCY_THRESHOLD and not args.force:
            logger.info(
                "DB already has %d interactions (threshold=%d). "
                "Use --force to re-seed. Exiting.",
                existing_count, IDEMPOTENCY_THRESHOLD,
            )
            return

        # ── Insert interactions ────────────────────────────────────
        logger.info("Inserting interactions...")
        results = seed_interactions(db, records, dry_run=False)

        db.commit()

        # Final count from DB
        final_count = db.scalar(
            select(func.count()).select_from(DrugInteraction)
        ) or 0

        logger.info("=" * 60)
        logger.info("  Inserted : %d", results["inserted"])
        logger.info("  Updated  : %d", results["updated"])
        logger.info("  Skipped  : %d", results["skipped"])
        logger.info("  Failed   : %d", results["failed"])
        logger.info("  DB total : %d interactions", final_count)
        logger.info("=" * 60)

        if results["failed"] > 0:
            logger.warning("%d interactions failed to import.", results["failed"])
        else:
            logger.info("Interactions seeding complete — no failures.")

    except Exception as exc:
        db.rollback()
        logger.exception("Interactions seeding failed: %s", exc)
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
