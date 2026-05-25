import csv
import io
import json

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.interaction import DrugInteraction
from app.models.medicine import ImportBatch
from app.schemas.medicine import MedicineCreate
from app.services.medicine_service import MedicineService
from app.services.text import normalize_key, split_values


class ImportService:
    def __init__(self, db: Session):
        self.db = db

    async def import_medicines(self, file: UploadFile, source_name: str, user_id: str) -> ImportBatch:
        rows = await self._read_rows(file)
        batch = ImportBatch(
            source_name=source_name,
            source_type="medicine_dataset",
            filename=file.filename,
            records_total=len(rows),
            created_by_user_id=user_id,
        )
        self.db.add(batch)
        self.db.flush()
        imported = 0
        errors: list[str] = []
        service = MedicineService(self.db)
        for index, row in enumerate(rows, start=1):
            try:
                generic_name = self._pick(row, "generic_name", "generic", "name")
                if not generic_name:
                    raise ValueError("generic_name is required")
                data = MedicineCreate(
                    generic_name=generic_name,
                    brand_name=self._pick(row, "brand_name", "brand"),
                    composition=self._pick(row, "composition", "ingredients"),
                    dosage_form=self._pick(row, "dosage_form", "form"),
                    strength=self._pick(row, "strength"),
                    side_effects=self._pick(row, "side_effects", "adverse_effects"),
                    precautions=self._pick(row, "precautions", "warnings"),
                    contraindications=self._pick(row, "contraindications"),
                    storage_instructions=self._pick(row, "storage_instructions", "storage"),
                    usage_guidelines=self._pick(row, "usage_guidelines", "usage", "directions"),
                    source_name=source_name,
                    source_url=self._pick(row, "source_url"),
                    source_version=self._pick(row, "source_version"),
                    source_date=self._pick(row, "source_date"),
                    aliases=split_values(self._pick(row, "aliases", "alias")),
                )
                service.create(data, import_batch_id=batch.id)
                imported += 1
            except Exception as exc:  # noqa: BLE001 - capture row-level import errors
                errors.append(f"row {index}: {exc}")
        batch.records_imported = imported
        batch.errors = "\n".join(errors) if errors else None
        self.db.commit()
        self.db.refresh(batch)
        return batch

    def list_batches(self, limit: int = 50, offset: int = 0) -> tuple[int, list[ImportBatch]]:
        total = self.db.scalar(select(func.count()).select_from(ImportBatch)) or 0
        rows = list(
            self.db.scalars(
                select(ImportBatch)
                .order_by(ImportBatch.created_at.desc())
                .limit(limit)
                .offset(offset)
            ).all()
        )
        return total, rows

    async def import_ddinter(self, file: UploadFile, source_name: str) -> dict[str, int]:
        rows = await self._read_rows(file)
        imported = 0
        for row in rows:
            drug_a = self._pick(row, "drug_a", "Drug_A", "name_a", "drug1")
            drug_b = self._pick(row, "drug_b", "Drug_B", "name_b", "drug2")
            if not drug_a or not drug_b:
                continue
            severity = self._map_severity(self._pick(row, "severity", "level", "risk_level"))

            # Canonical key ordering: drug_a_key < drug_b_key lexicographically
            key_a = normalize_key(drug_a)
            key_b = normalize_key(drug_b)
            if key_a > key_b:
                key_a, key_b = key_b, key_a
                drug_a, drug_b = drug_b, drug_a

            update_fields = {
                "drug_a_name": drug_a,
                "drug_b_name": drug_b,
                "severity": severity,
                "explanation": self._pick(row, "explanation", "description", "interaction") or "Interaction noted.",
                "mechanism": self._pick(row, "mechanism"),
                "recommendations": self._pick(row, "recommendations", "management", "precaution"),
                "source_name": source_name,
                "source_url": self._pick(row, "source_url"),
                "source_version": self._pick(row, "source_version"),
                "confidence": 0.8,
            }

            # Use a SAVEPOINT to handle the race condition where two concurrent
            # imports try to insert the same (drug_a_key, drug_b_key) pair.
            # begin_nested() creates a SAVEPOINT; IntegrityError rolls it back
            # without aborting the outer transaction, then we fall through to
            # update the existing row.
            try:
                with self.db.begin_nested():
                    self.db.add(
                        DrugInteraction(drug_a_key=key_a, drug_b_key=key_b, **update_fields)
                    )
            except IntegrityError:
                # Duplicate pair — update existing record in-place
                existing = self.db.scalar(
                    select(DrugInteraction).where(
                        DrugInteraction.drug_a_key == key_a,
                        DrugInteraction.drug_b_key == key_b,
                    )
                )
                if existing:
                    for field, value in update_fields.items():
                        setattr(existing, field, value)

            imported += 1

        self.db.commit()
        return {"records_total": len(rows), "records_imported": imported}

    async def _read_rows(self, file: UploadFile) -> list[dict[str, str]]:
        # Fix 1: Enforce upload size limit.  Read one extra byte so we can
        # detect oversized files without loading the entire payload into RAM.
        content = await file.read(settings.max_upload_bytes + 1)
        if len(content) > settings.max_upload_bytes:
            limit_mb = settings.max_upload_bytes // (1024 * 1024)
            raise HTTPException(
                status_code=status.HTTP_413_CONTENT_TOO_LARGE,
                detail=f"Upload exceeds the {limit_mb} MB size limit.",
            )

        filename = (file.filename or "").lower()

        if filename.endswith(".json"):
            data = json.loads(content.decode("utf-8"))
            if not isinstance(data, list):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="JSON must be a list")
            return [dict(row) for row in data]

        if filename.endswith(".xlsx"):
            # Fix 4: openpyxl raises InvalidFileException (and other errors)
            # for corrupt or truncated workbooks — surface this as a 400 so
            # the admin gets a clear message instead of an unhandled 500.
            try:
                workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            except (InvalidFileException, Exception) as exc:  # noqa: BLE001
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid or corrupt XLSX file. Ensure the file is a valid Excel workbook.",
                ) from exc
            sheet = workbook.active
            if sheet is None:
                return []
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            return [
                {headers[index]: "" if value is None else str(value) for index, value in enumerate(row)}
                for row in rows[1:]
            ]

        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            return list(csv.DictReader(io.StringIO(text)))

        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Use CSV, JSON, or XLSX")

    def _pick(self, row: dict, *keys: str) -> str | None:
        lowered = {str(key).casefold(): value for key, value in row.items()}
        for key in keys:
            value = lowered.get(key.casefold())
            if value is not None and str(value).strip():
                return str(value).strip()
        return None

    def _map_severity(self, value: str | None) -> str:
        normalized = normalize_key(value)
        if normalized in {"dangerous", "major", "high", "contraindicated", "serious"}:
            return "dangerous"
        if normalized in {"moderate", "medium", "monitor", "caution"}:
            return "moderate"
        if normalized in {"safe", "none", "minor", "low"}:
            return "safe"
        return "unknown"
